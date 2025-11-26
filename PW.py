import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import statistics as stats
from openpyxl.styles import PatternFill
# Pacchetti per pop-up interattivi
import tkinter as tk
from tkinter import simpledialog



# ====================================================
# 1. CARICAMENTO E PULIZIA DATI
# ====================================================

df = pd.read_excel(
    "C:/Users/dalil/OneDrive/Documenti/2° anno/PYTHON/Online_Retail.xlsx"
)

# Copia del dataset
df_clean = df.copy()

# Rimozione valori nulli nelle colonne chiave
df_clean = df_clean.dropna(
    subset=[
        "InvoiceNo",
        "StockCode",
        "Description",
        "Quantity",
        "InvoiceDate",
        "UnitPrice",
        "CustomerID",
        "Country",
    ]
)

# Eliminazione quantità negative (resi)
df_clean = df_clean[df_clean["Quantity"] > 0]

# Eliminazione duplicati
df_clean = df_clean.drop_duplicates()

# Creazione colonna totale vendita
df_clean["TotalSales"] = df_clean["Quantity"] * df_clean["UnitPrice"]

# Conversione a datetime e colonne temporali
df_clean["InvoiceDate"] = pd.to_datetime(df_clean["InvoiceDate"])
df_clean["YearWeek"] = df_clean["InvoiceDate"].dt.strftime("%Y-%U")
df_clean["YearMonth"] = df_clean["InvoiceDate"].dt.to_period("M")

# ====================================================
# POP-UP: CREAZIONE FINESTRA NASCOSTA TKINTER
# ====================================================

root = tk.Tk()
root.withdraw()  # nasconde la finestra principale

# Pop-up per i titoli dei grafici
title_week = simpledialog.askstring(
    "Titolo grafico",
    "Inserisci un titolo per il grafico delle vendite settimanali:"
)
title_month = simpledialog.askstring(
    "Titolo grafico",
    "Inserisci un titolo per il grafico delle vendite mensili:"
)
title_top = simpledialog.askstring(
    "Titolo grafico",
    "Inserisci un titolo per il grafico delle vendite mensili dei top prodotti:"
)
title_corr = simpledialog.askstring(
    "Titolo grafico",
    "Inserisci un titolo per la heatmap di correlazione fra prodotti:"
)

# ====================================================
# 2. ANALISI TEMPORALE: VENDITE SETTIMANALI E MENSILI
# ====================================================

sales_week = df_clean.groupby("YearWeek")["TotalSales"].sum().sort_index()
sales_month = df_clean.groupby("YearMonth")["TotalSales"].sum().sort_index()

# Grafico vendite settimanali
plt.figure(figsize=(12, 5))
sales_week.plot()
plt.title(title_week if title_week else "Vendite Settimanali")
plt.ylabel("Totale vendite")
plt.tight_layout()
plt.show()

# Grafico vendite mensili
plt.figure(figsize=(12, 5))
sales_month.plot()
plt.title(title_month if title_month else "Vendite Mensili")
plt.ylabel("Totale vendite")
plt.tight_layout()
plt.show()


# ====================================================
# 3. ANALISI STAGIONALE PER TOP PRODOTTI
# ====================================================

top_products = (
    df_clean.groupby("Description")["TotalSales"]
    .sum()
    .sort_values(ascending=False)
    .head(10)
)

df_top = df_clean[df_clean["Description"].isin(top_products.index)]

pivot = df_top.pivot_table(
    values="TotalSales",
    index="YearMonth",
    columns="Description",
    aggfunc="sum",
)

plt.figure(figsize=(15, 7))
pivot.plot(ax=plt.gca())
plt.title(title_top if title_top else "Vendite mensili per top prodotti")
plt.ylabel("Totale vendite")
plt.ylim(0, 25000)
plt.tight_layout()
plt.show()


# ====================================================
# 4. DEVIAZIONE STANDARD PREZZO ED ELASTICITÀ DOMANDA
# ====================================================

df_el = df_clean[
    ["InvoiceNo", "StockCode", "Description", "Quantity", "UnitPrice", "InvoiceDate", "TotalSales"]
].copy()

df_el["InvoiceDate"] = pd.to_datetime(df_el["InvoiceDate"])
df_el["YearMonth"] = df_el["InvoiceDate"].dt.to_period("M")

# Deviazione standard del prezzo per StockCode

def std_unit_price(group):
    """Calcola la deviazione standard (sample) dei prezzi usando statistics.stdev."""
    values = group["UnitPrice"].tolist()
    if len(values) > 1:
        return stats.stdev(values)   # deviazione standard campionaria
    else:
        return 0.0                   

price_var = (
    df_el.groupby("StockCode")
    .apply(std_unit_price)
    .reset_index()
    .rename(columns={0: "StdDev_UnitPrice"})
)

# Ordinamento per variabilità di prezzo
price_var = price_var.sort_values(by="StdDev_UnitPrice", ascending=False)


def price_comment(std):
    if std > 100:
        return "Variazione estremamente elevata"
    elif std > 50:
        return "Variazione molto alta"
    elif std > 20:
        return "Variazione significativa"
    elif std > 10:
        return "Variazione moderata"
    else:
        return "Variazione bassa"


price_var["Commento"] = price_var["StdDev_UnitPrice"].apply(price_comment)

# TOTALE DELLE VENDITE PER PRODOTTO
sales_per_product = (
    df_clean.groupby("StockCode")["TotalSales"]
    .sum()
    .reset_index()
    .rename(columns={"TotalSales": "TotalSales_Product"})
)


price_var = price_var.merge(sales_per_product, on="StockCode", how="left")

print("Tabella deviazione standard prezzo (statistics) + commento + totale vendite:")
print(price_var.head(10))

# Elasticità sui top 10 per variabilità di prezzo

selected_products = price_var.head(10)["StockCode"].tolist()

elasticity_tables = {}

for code in selected_products:
    temp = df_el[df_el["StockCode"] == code]

    monthly = temp.groupby("YearMonth").agg(
        AvgPrice=("UnitPrice", "mean"),
        TotalQty=("Quantity", "sum"),
    )

    monthly["ΔPrice"] = monthly["AvgPrice"].pct_change()
    monthly["ΔQty"] = monthly["TotalQty"].pct_change()
    monthly["Elasticity"] = monthly["ΔQty"] / monthly["ΔPrice"]

    elasticity_tables[code] = monthly

# Interpretazione elasticità
interpretation = {}

for code, table in elasticity_tables.items():
    mean_el = table["Elasticity"].mean()

    if pd.isna(mean_el):
        status = "Dati insufficienti per calcolare elasticità"
    elif mean_el < -1:
        status = "Domanda elastica (alta sensibilità al prezzo)"
    elif mean_el < 0:
        status = "Domanda anelastica (bassa sensibilità al prezzo)"
    else:
        status = "Elasticità positiva (anomalia o dinamiche promozionali)"

    interpretation[code] = {
        "Elasticità media": mean_el,
        "Valutazione": status,
    }

elasticity_df = pd.DataFrame.from_dict(interpretation, orient="index")
elasticity_df.index.name = "StockCode"

print("\nSintesi elasticità media per i top 10 prodotti:")
print(elasticity_df)


# ====================================================
# 5. CORRELAZIONE TRA PRODOTTI (TOP 20, HEATMAP)
# ====================================================

top20_codes = df_clean["StockCode"].value_counts().head(20).index
df_corr = df_clean[df_clean["StockCode"].isin(top20_codes)].copy()

basket = df_corr.pivot_table(
    index="InvoiceNo",
    columns="StockCode",
    values="Quantity",
    aggfunc="sum",
).fillna(0)

corr_matrix = basket.corr()

plt.figure(figsize=(12, 10))
sns.heatmap(corr_matrix, cmap="coolwarm")
plt.title(title_corr if title_corr else "Correlazione fra prodotti (Top 20)")
plt.tight_layout()
plt.show()


# ====================================================
# 6. SALVATAGGIO GRAFICI SU FILE PNG
# ====================================================

# Grafico vendite settimanali
fig1, ax1 = plt.subplots(figsize=(8, 4))
sales_week.plot(ax=ax1)
ax1.set_title(title_week if title_week else "Vendite Settimanali")
ax1.set_ylabel("Totale vendite")
fig1.tight_layout()
fig1.savefig("grafico_vendite_settimanali.png", dpi=150)

# Grafico vendite mensili
fig2, ax2 = plt.subplots(figsize=(8, 4))
sales_month.plot(ax=ax2)
ax2.set_title(title_month if title_month else "Vendite Mensili")
ax2.set_ylabel("Totale vendite")
fig2.tight_layout()
fig2.savefig("grafico_vendite_mensili.png", dpi=150)

# Grafico vendite mensili per top prodotti
fig3, ax3 = plt.subplots(figsize=(10, 5))
pivot.plot(ax=ax3)
ax3.set_title(title_top if title_top else "Vendite mensili per top prodotti")
ax3.set_ylabel("Totale vendite")
ax3.set_ylim(0, 25000)
fig3.tight_layout()
fig3.savefig("grafico_top_prodotti_mensili.png", dpi=150)

# Heatmap correlazioni top 20 prodotti
fig4, ax4 = plt.subplots(figsize=(8, 6))
sns.heatmap(corr_matrix, cmap="coolwarm", ax=ax4)
ax4.set_title(title_corr if title_corr else "Correlazione fra prodotti (Top 20)")
fig4.tight_layout()
fig4.savefig("heatmap_correlazioni_prodotti.png", dpi=150)

plt.close("all")


# ====================================================
# 7. CREAZIONE FILE EXCEL
# ====================================================

sales_week_df = sales_week.to_frame(name="TotalSales")
sales_month_df = sales_month.to_frame(name="TotalSales")

output_file = "Analisi_Online_Retail.xlsx"

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    # --- Fogli con i dati ---
    sales_week_df.to_excel(writer, sheet_name="Vendite_settimanali")
    sales_month_df.to_excel(writer, sheet_name="Vendite_mensili")
    pivot.to_excel(writer, sheet_name="Top_prodotti_mese")
    price_var.to_excel(
        writer,
        sheet_name="Var_prezzo_prodotti",   # StdDev (statistics), Commento, Tot vendite
        index=False,
    )
    corr_matrix.to_excel(writer, sheet_name="Correlazione_prodotti")
    elasticity_df.to_excel(writer, sheet_name="Elasticita_domanda")

    workbook = writer.book
    ws_week = workbook["Vendite_settimanali"]
    ws_month = workbook["Vendite_mensili"]
    ws_top = workbook["Top_prodotti_mese"]
    ws_corr = workbook["Correlazione_prodotti"]

    # ===== COLORAZIONE DIAGONALE CORRELAZIONE (VALORE = 1) =====
    red_fill = PatternFill(start_color="FFFF0000",
                           end_color="FFFF0000",
                           fill_type="solid")

    n = len(corr_matrix.index)

    for i in range(n):
        row = 2 + i   
        col = 2 + i   
        cell = ws_corr.cell(row=row, column=col)

        # Controllo per verificare che il valore della cella sia 1
        if cell.value == 1 or cell.value == 1.0:
            cell.fill = red_fill


    # --- 1) Grafico vendite settimanali nel foglio "Vendite_settimanali" ---
    img1 = Image("grafico_vendite_settimanali.png")
    img1.width = 480
    img1.height = 280

    last_col_week = ws_week.max_column          # ultima colonna con dati
    target_col_week = last_col_week + 2         # +1 colonna vuota
    col_letter_week = get_column_letter(target_col_week)
    anchor_week = f"{col_letter_week}2"         # riga 2 per non sovrapporsi all’intestazione

    ws_week.add_image(img1, anchor_week)

    # --- 2) Grafico vendite mensili nel foglio "Vendite_mensili" ---
    img2 = Image("grafico_vendite_mensili.png")
    img2.width = 480
    img2.height = 280

    last_col_month = ws_month.max_column
    target_col_month = last_col_month + 2
    col_letter_month = get_column_letter(target_col_month)
    anchor_month = f"{col_letter_month}2"

    ws_month.add_image(img2, anchor_month)

    # --- 3) Grafico top prodotti mensili nel foglio "Top_prodotti_mese" ---
    img3 = Image("grafico_top_prodotti_mensili.png")
    img3.width = 520
    img3.height = 300

    last_col_top = ws_top.max_column
    target_col_top = last_col_top + 2
    col_letter_top = get_column_letter(target_col_top)
    anchor_top = f"{col_letter_top}2"

    ws_top.add_image(img3, anchor_top)

    # --- 4) Heatmap correlazioni nel foglio "Correlazione_prodotti" ---
    img4 = Image("heatmap_correlazioni_prodotti.png")
    img4.width = 480
    img4.height = 320

    last_col_corr = ws_corr.max_column
    target_col_corr = last_col_corr + 2
    col_letter_corr = get_column_letter(target_col_corr)
    anchor_corr = f"{col_letter_corr}2"

    ws_corr.add_image(img4, anchor_corr)

print("File Excel creato correttamente:", output_file)

# Chiudiamo la root di Tkinter
root.destroy()
